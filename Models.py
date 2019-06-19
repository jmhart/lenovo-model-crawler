from openpyxl import Workbook, load_workbook


class Models:
    def __init__(self):
        self._model_ids = []
        self.load_models()

    @property
    def model_ids(self):
        return self._model_ids

    def load_models(self):
        wb = load_workbook("models.xlsx")
        ws = wb["Sheet1"]
        self._model_ids = [i.value for i in ws["a"]]
        wb.close

    def write_model_names(self, model_names):
        wb = Workbook()
        ws = wb.active

        for model_name in model_names:
            ws.append(model_name)

        wb.save("new_models.xlsx")
